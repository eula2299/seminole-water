#!/usr/bin/env python3
import json,hashlib,re
from pathlib import Path
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]
SRC=ROOT/'raw'/'Chem_Report_2024.xlsx'; OUT=ROOT/'data'/'all_contaminant_records.json'
wb=load_workbook(SRC,read_only=True,data_only=True); ws=wb[wb.sheetnames[0]]
rows=ws.iter_rows(values_only=True); hdr=[str(x) for x in next(rows)]
out=[]
for vals in rows:
 d=dict(zip(hdr,vals)); p=str(d.get('PWSID') or '')
 if not p.startswith('359'): continue
 desc=str(d.get('CONTAMDESC') or '').strip(); group=str(d.get('CONTAMGROUP') or '').strip()
 raw=d.get('RESULTS'); unit=d.get('UOM')
 rec={
  'pwsid':p,'system_name':d.get('MAILINGNAME'),'system_type':d.get('SYSTEMTYPE'),'population':d.get('POPULATION'),
  'contaminant_group':group,'analyte_code':str(d.get('CONTAMCODE') or ''),'analyte':desc,
  'sample_type':d.get('SAMPLETYPE'),'result':raw,'unit':unit,'mcl':d.get('MCL'),'sample_date':d.get('SAMPLEDATE'),
  'sample_timestamp':d.get('SAMPTIMESTAMP'),'location_code':d.get('LOCATIONCODE'),'entry_point':d.get('ENTRYPOINT'),
  'lab_id':d.get('LAB'),'method_id':d.get('METHOD'),'analysis_date':d.get('ANALYSISDATE'),'mdl':d.get('MDL'),'rdl':d.get('RDL'),
  'sample_id':str(d.get('SAMPLENUMBER') or ''),'remarks':d.get('REMARKS'),'source_id':'fdep-chemical-2024','publisher':'Florida DEP'
 }
 fp='|'.join(str(rec.get(k) or '') for k in ['pwsid','sample_id','analyte_code','sample_date','location_code','lab_id','method_id','result','unit'])
 rec['record_fingerprint']=hashlib.sha256(fp.encode()).hexdigest()
 out.append(rec)
OUT.write_text(json.dumps(out,indent=2,default=str))
summary={}
for r in out: summary[r['contaminant_group']]=summary.get(r['contaminant_group'],0)+1
(ROOT/'data'/'all_contaminant_summary.json').write_text(json.dumps({'records':len(out),'groups':summary,'systems':len(set(r['pwsid'] for r in out)),'analytes':len(set(r['analyte'] for r in out))},indent=2))
print(json.dumps({'records':len(out),'groups':summary},indent=2))
