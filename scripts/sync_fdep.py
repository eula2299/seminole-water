#!/usr/bin/env python3
"""Download current FDEP workbook, filter all Seminole County PWSIDs (county code 359), and stage changes for review.
Never auto-publishes: writes data/pending_records.json and data/pending_report.json.
"""
from pathlib import Path
import urllib.request, json, hashlib, shutil, datetime
from openpyxl import load_workbook
ROOT=Path(__file__).resolve().parents[1]; RAW=ROOT/'raw'; DATA=ROOT/'data'; RAW.mkdir(exist_ok=True)
URL='https://floridadep.gov/sites/default/files/Chem_Report_2024.xlsx'
out=RAW/'latest.xlsx'; urllib.request.urlretrieve(URL,out)
metals=['ANTIMONY','ARSENIC','BARIUM','BERYLLIUM','CADMIUM','CHROMIUM','COPPER','LEAD','MERCURY','SELENIUM','THALLIUM','NICKEL','ALUMINUM','IRON','MANGANESE','SILVER','ZINC']
wb=load_workbook(out,read_only=True,data_only=True); ws=wb.active; it=ws.iter_rows(values_only=True); h=list(next(it)); ix={x:i for i,x in enumerate(h)}; rows=[]
for r in it:
 p=str(r[ix['PWSID']] or '')
 if not p.startswith('359'): continue
 desc=str(r[ix['CONTAMDESC']] or '').upper(); metal=next((m for m in metals if m in desc),None)
 if not metal: continue
 rows.append({'pwsid':p,'system_name':r[ix['MAILINGNAME']],'system_type':r[ix['SYSTEMTYPE']],'population':r[ix['POPULATION']],'metal':metal,'contaminant':r[ix['CONTAMDESC']],'contam_code':r[ix['CONTAMCODE']],'sample_type':r[ix['SAMPLETYPE']],'result':r[ix['RESULTS']],'unit':r[ix['UOM']],'mcl':r[ix['MCL']],'sample_date':r[ix['SAMPLEDATE']],'reported_at':r[ix['SAMPTIMESTAMP']],'source_year':2024,'source_file':out.name,'source_level':'public-water-system/compliance sample','detected':isinstance(r[ix['RESULTS']],(int,float)) and r[ix['RESULTS']]>0})
(DATA/'pending_records.json').write_text(json.dumps(rows,indent=2,default=str)); old=json.loads((DATA/'metal_records.json').read_text())
report={'checked_at':datetime.datetime.now(datetime.timezone.utc).isoformat(),'source':URL,'old_rows':len(old),'new_rows':len(rows),'changed':hashlib.sha256(json.dumps(old,sort_keys=True,default=str).encode()).hexdigest()!=hashlib.sha256(json.dumps(rows,sort_keys=True,default=str).encode()).hexdigest(),'status':'REVIEW REQUIRED; run scripts/approve_pending.py after validation'}
(DATA/'pending_report.json').write_text(json.dumps(report,indent=2));print(json.dumps(report,indent=2))
